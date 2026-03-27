"""
Workpaper generator – produces an Excel file suitable for audit substantiation.

Sheets:
    1. Summary          – period, entity, rates used (with source URLs)
    2. BS Translation   – AUD balances × rates = USD balances + movement
    3. IS Translation   – AUD amounts × average rate = USD amounts
    4. Exch Adjustment  – calculation of the exchange plug
    5. Journal Entry    – complete JE in human-readable format
"""
import io
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
_ARIAL = "Arial"

def _font(bold=False, size=10, color="000000", italic=False):
    return Font(name=_ARIAL, size=size, bold=bold, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# Colour palette
_C_DARK_BLUE  = "1F4E79"   # header background
_C_MID_BLUE   = "BDD7EE"   # sub-header background
_C_LIGHT_BLUE = "DEEAF1"   # section label background
_C_GREEN      = "E2EFDA"   # total row background
_C_YELLOW     = "FFFF00"   # manual input highlight
_C_WHITE      = "FFFFFF"

# Number formats
_FMT_USD   = '#,##0.00;(#,##0.00);"-"'
_FMT_RATE  = "0.000000"
_FMT_DATE  = "MM/DD/YYYY"

_THIN = _border("thin")


# ---------------------------------------------------------------------------
# Low-level cell helpers
# ---------------------------------------------------------------------------

def _set(ws, row, col, value, font=None, fill=None, alignment=None,
         number_format=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:        cell.font = font
    if fill:        cell.fill = fill
    if alignment:   cell.alignment = alignment
    if number_format: cell.number_format = number_format
    if border:      cell.border = border
    return cell


def _header(ws, row, col, text, span=1, row_span=1):
    """Dark-blue header cell, white bold text, centered."""
    cell = _set(
        ws, row, col, text,
        font=_font(bold=True, color=_C_WHITE),
        fill=_fill(_C_DARK_BLUE),
        alignment=_align("center", "center", wrap=True),
        border=_THIN,
    )
    if span > 1 or row_span > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row + row_span - 1, end_column=col + span - 1,
        )
    return cell


def _sub_header(ws, row, col, text, span=1):
    cell = _set(
        ws, row, col, text,
        font=_font(bold=True),
        fill=_fill(_C_MID_BLUE),
        alignment=_align("center"),
        border=_THIN,
    )
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return cell


def _total_row(ws, row, first_col, last_col, label="TOTAL", label_col=None):
    if label_col:
        _set(ws, row, label_col, label,
             font=_font(bold=True), fill=_fill(_C_GREEN), border=_THIN)
    for c in range(first_col, last_col + 1):
        if label_col and c == label_col:
            continue
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(_C_GREEN)
        cell.font = _font(bold=True)
        cell.border = _THIN


def _col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_summary(ws, period_date: date, entity_name: str, rates: dict,
                   je_summary: dict):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    _col_width(ws, 1, 32)
    _col_width(ws, 2, 22)
    _col_width(ws, 3, 50)

    # Title
    _set(ws, 1, 1, "AUS → USD TRANSLATION WORKPAPER",
         font=_font(bold=True, size=14), alignment=_align("left"))
    ws.merge_cells("A1:C1")

    _set(ws, 2, 1, entity_name or "Australian Subsidiary",
         font=_font(size=11, italic=True))
    ws.merge_cells("A2:C2")

    _set(ws, 3, 1, f"Period ended: {period_date.strftime('%B %d, %Y')}",
         font=_font(size=11))
    ws.merge_cells("A3:C3")

    _set(ws, 4, 1, f"Prepared: {date.today().strftime('%B %d, %Y')}",
         font=_font(size=10, italic=True))
    ws.merge_cells("A4:C4")

    # Section: Exchange Rates
    row = 6
    _header(ws, row, 1, "EXCHANGE RATES USED", span=3)
    row += 1
    for label, col in [("Description", 1), ("Rate (AUD/USD)", 2), ("Source / Notes", 3)]:
        _sub_header(ws, row, col, label)
    row += 1

    rate_data = [
        (
            f"Current period end  ({rates['current_actual_date']})",
            rates["current_rate"],
            f"Frankfurter API  |  {rates['current_url']}",
        ),
        (
            f"Prior period end  ({rates['prior_actual_date']})",
            rates["prior_rate"],
            f"Frankfurter API  |  {rates['prior_url']}",
        ),
        (
            "Average rate (IS translation)",
            rates["avg_rate"],
            "= (Current + Prior) / 2",
        ),
    ]
    for desc, rate_val, note in rate_data:
        _set(ws, row, 1, desc, font=_font(), border=_THIN)
        _set(ws, row, 2, rate_val, font=_font(color="0000FF"),
             number_format=_FMT_RATE, border=_THIN,
             alignment=_align("right"))
        _set(ws, row, 3, note, font=_font(italic=True), border=_THIN)
        row += 1

    # Section: JE Balance Check
    row += 1
    _header(ws, row, 1, "JOURNAL ENTRY BALANCE CHECK", span=3)
    row += 1
    _sub_header(ws, row, 1, "Item")
    _sub_header(ws, row, 2, "Amount (USD)")
    _sub_header(ws, row, 3, "Status")
    row += 1

    balanced_str = "✓  BALANCED" if je_summary["balanced"] else "✗  OUT OF BALANCE"
    bal_color = "00B050" if je_summary["balanced"] else "FF0000"
    for label, val in [("Total Debits", je_summary["total_debits"]),
                       ("Total Credits", je_summary["total_credits"])]:
        _set(ws, row, 1, label, font=_font(), border=_THIN)
        _set(ws, row, 2, val, font=_font(), number_format=_FMT_USD,
             border=_THIN, alignment=_align("right"))
        _set(ws, row, 3, "", border=_THIN)
        row += 1
    _set(ws, row, 1, "Difference", font=_font(bold=True), border=_THIN)
    _set(ws, row, 2, je_summary["difference"], font=_font(bold=True),
         number_format=_FMT_USD, border=_THIN, alignment=_align("right"))
    _set(ws, row, 3, balanced_str, font=_font(bold=True, color=bal_color), border=_THIN)


# ---------------------------------------------------------------------------

def _build_bs_translation(ws, translated_bs: pd.DataFrame,
                           current_col_label: str, prior_col_label: str):
    ws.title = "BS Translation"
    ws.sheet_view.showGridLines = False

    # Column widths
    widths = [32, 12, 12, 12, 12, 14, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        _col_width(ws, i, w)

    # Header
    _header(ws, 1, 1, "BALANCE SHEET TRANSLATION (ASC 830 – Period-End Rate)", span=9)

    # Sub-headers row 2
    headers = [
        "Xero Account", "Type",
        f"AUD Balance\n{current_col_label}", f"AUD Balance\n{prior_col_label}",
        "Rate\n(Current)", "Rate\n(Prior)",
        "USD Balance\n(Current)", "USD Balance\n(Prior)",
        "USD Movement\n(JE Amount)",
    ]
    for c, h in enumerate(headers, 1):
        _sub_header(ws, 2, c, h)

    row = 3
    last_section = None
    total_movement = 0.0

    for _, r in translated_bs.iterrows():
        # Section separator
        section = r["xero_type"]
        if section != last_section:
            _set(ws, row, 1, section.upper(),
                 font=_font(bold=True), fill=_fill(_C_LIGHT_BLUE), border=_THIN)
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=9)
            row += 1
            last_section = section

        unmapped = r["qb_account"] == "*** UNMAPPED ***"
        txt_color = "FF0000" if unmapped else "000000"
        label = f"{r['xero_account']}  →  {r['qb_account']}"

        vals = [
            label, r["je_section"],
            r["current_aud"], r["prior_aud"],
            r["current_rate"], r["prior_rate"],
            r["current_usd"], r["prior_usd"],
            r["movement_usd"],
        ]
        for c, v in enumerate(vals, 1):
            fmt = _FMT_RATE if c in (5, 6) else (_FMT_USD if c >= 3 else None)
            _set(ws, row, c, v,
                 font=_font(color=txt_color),
                 number_format=fmt,
                 alignment=_align("right" if c >= 3 else "left"),
                 border=_THIN)
        total_movement += r["movement_usd"] or 0
        row += 1

    # Total row
    _total_row(ws, row, 1, 9, label_col=1)
    _set(ws, row, 1, "NET BS MOVEMENT (USD)",
         font=_font(bold=True), fill=_fill(_C_GREEN), border=_THIN)
    _set(ws, row, 9, total_movement,
         font=_font(bold=True), number_format=_FMT_USD,
         fill=_fill(_C_GREEN), border=_THIN, alignment=_align("right"))


# ---------------------------------------------------------------------------

def _build_is_translation(ws, translated_is: pd.DataFrame, period_label: str):
    ws.title = "IS Translation"
    ws.sheet_view.showGridLines = False

    widths = [40, 12, 14, 12, 14]
    for i, w in enumerate(widths, 1):
        _col_width(ws, i, w)

    _header(ws, 1, 1, "INCOME STATEMENT TRANSLATION (ASC 830 – Average Rate)", span=5)

    headers = [
        "Xero Account  →  QB Account",
        "QB Type",
        f"AUD Amount\n({period_label})",
        "Avg Rate",
        "USD Amount",
    ]
    for c, h in enumerate(headers, 1):
        _sub_header(ws, 2, c, h)

    row = 3
    total_usd = 0.0
    for _, r in translated_is.iterrows():
        unmapped = r["qb_account"] == "*** UNMAPPED ***"
        txt_color = "FF0000" if unmapped else "000000"
        label = f"{r['xero_account']}  →  {r['qb_account']}"

        vals = [label, r["qb_type"], r["aud_amount"], r["avg_rate"], r["usd_amount"]]
        for c, v in enumerate(vals, 1):
            fmt = _FMT_RATE if c == 4 else (_FMT_USD if c in (3, 5) else None)
            _set(ws, row, c, v,
                 font=_font(color=txt_color),
                 number_format=fmt,
                 alignment=_align("right" if c >= 3 else "left"),
                 border=_THIN)
        total_usd += r["usd_amount"] or 0
        row += 1

    _total_row(ws, row, 1, 5, label_col=1)
    _set(ws, row, 1, "NET IS ACTIVITY (USD)",
         font=_font(bold=True), fill=_fill(_C_GREEN), border=_THIN)
    _set(ws, row, 5, total_usd,
         font=_font(bold=True), number_format=_FMT_USD,
         fill=_fill(_C_GREEN), border=_THIN, alignment=_align("right"))


# ---------------------------------------------------------------------------

def _build_exch_adj(ws, je_lines: list[dict], rates: dict):
    ws.title = "Exch Adjustment"
    ws.sheet_view.showGridLines = False
    _col_width(ws, 1, 45)
    _col_width(ws, 2, 18)
    _col_width(ws, 3, 40)

    _header(ws, 1, 1, "EXCHANGE ADJUSTMENT CALCULATION", span=3)

    row = 3
    _set(ws, row, 1, "METHODOLOGY", font=_font(bold=True, size=11))
    row += 1
    note = (
        "Under the simplified ASC 830 approach used here, the Balance Sheet is translated "
        "at the period-end spot rate and the Income Statement is translated at the average rate "
        "(arithmetic mean of current and prior period-end rates).  Because these rates differ, "
        "a plug is required to make the combined Journal Entry balance.  This plug is recorded "
        "as a Currency Exchange Gain / (Loss) in the Income Statement."
    )
    _set(ws, row, 1, note, font=_font(italic=True), alignment=_align(wrap=True))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.row_dimensions[row].height = 60
    row += 2

    # Rates used
    _sub_header(ws, row, 1, "Rate Used", span=1)
    _sub_header(ws, row, 2, "Value")
    _sub_header(ws, row, 3, "Description")
    row += 1
    rate_rows = [
        ("Current period-end rate", rates["current_rate"],
         f"AUD/USD as of {rates['current_actual_date']}"),
        ("Prior period-end rate",   rates["prior_rate"],
         f"AUD/USD as of {rates['prior_actual_date']}"),
        ("Average rate (IS)",       rates["avg_rate"],
         "(Current + Prior) ÷ 2"),
    ]
    for lbl, val, desc in rate_rows:
        _set(ws, row, 1, lbl, font=_font(), border=_THIN)
        _set(ws, row, 2, val, font=_font(color="0000FF"), number_format=_FMT_RATE,
             border=_THIN, alignment=_align("right"))
        _set(ws, row, 3, desc, font=_font(italic=True), border=_THIN)
        row += 1

    row += 1

    # Running totals
    _sub_header(ws, row, 1, "Component", span=1)
    _sub_header(ws, row, 2, "USD")
    _sub_header(ws, row, 3, "Notes")
    row += 1

    bs_lines  = [l for l in je_lines if l["section"] == "BS"]
    is_lines  = [l for l in je_lines if l["section"] == "IS"]
    ex_lines  = [l for l in je_lines if l["section"] == "EXCH"]

    def net(lines):
        return sum((l["debit"] or 0) - (l["credit"] or 0) for l in lines)

    bs_net = net(bs_lines)
    is_net = net(is_lines)
    ex_net = net(ex_lines)

    exch_adj = 0.0
    for l in ex_lines:
        if l["account_name"] == "Currency Exchange Gain (Loss)":
            exch_adj += (l["credit"] or 0) - (l["debit"] or 0)

    component_rows = [
        ("BS section net (debits − credits)", bs_net,
         "Non-equity BS movements at period-end rate"),
        ("IS section net (debits − credits)", is_net,
         "IS activity + equity BS movements at average/current rate"),
        ("Exchange adjustment", exch_adj,
         "Plug to balance; positive = Gain, negative = Loss"),
    ]
    for lbl, val, desc in component_rows:
        _set(ws, row, 1, lbl, font=_font(), border=_THIN)
        _set(ws, row, 2, val, font=_font(), number_format=_FMT_USD,
             border=_THIN, alignment=_align("right"))
        _set(ws, row, 3, desc, font=_font(italic=True), border=_THIN)
        row += 1

    total = bs_net + is_net + ex_net
    _total_row(ws, row, 1, 3, label_col=1)
    _set(ws, row, 1, "Sum (should be zero if JE balanced)",
         font=_font(bold=True), fill=_fill(_C_GREEN), border=_THIN)
    color = "00B050" if abs(total) < 0.02 else "FF0000"
    _set(ws, row, 2, total, font=_font(bold=True, color=color),
         number_format=_FMT_USD, fill=_fill(_C_GREEN),
         border=_THIN, alignment=_align("right"))


# ---------------------------------------------------------------------------

def _build_je_sheet(ws, je_lines: list[dict], je_summary_data: dict):
    ws.title = "Journal Entry"
    ws.sheet_view.showGridLines = False

    widths = [6, 12, 16, 14, 12, 55, 10, 35, 14, 14]
    for i, w in enumerate(widths, 1):
        _col_width(ws, i, w)

    _header(ws, 1, 1, "JOURNAL ENTRY – COMPLETE LISTING", span=10)

    col_labels = [
        "#", "Date", "Type", "Num", "Name",
        "Memo / Description", "Acct #", "Account Name", "Debit", "Credit",
    ]
    for c, h in enumerate(col_labels, 1):
        _sub_header(ws, 2, c, h)

    row = 3
    last_section = None
    section_labels = {"BS": "BALANCE SHEET ACCOUNTS",
                      "IS": "INCOME STATEMENT ACCOUNTS",
                      "EXCH": "EXCHANGE ADJUSTMENT"}

    for i, ln in enumerate(je_lines, 1):
        sec = ln["section"]
        if sec != last_section:
            _set(ws, row, 1, section_labels.get(sec, sec),
                 font=_font(bold=True), fill=_fill(_C_LIGHT_BLUE), border=_THIN)
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=10)
            row += 1
            last_section = sec

        unmapped = not ln.get("is_mapped", True)
        txt_color = "FF0000" if unmapped else "000000"

        vals = [
            i,
            ln["date"],
            ln["type"],
            ln["num"],
            ln["name"],
            ln["memo"],
            ln["account_number"],
            ln["account_name"],
            ln["debit"],
            ln["credit"],
        ]
        for c, v in enumerate(vals, 1):
            fmt = _FMT_USD if c in (9, 10) else None
            _set(ws, row, c, v,
                 font=_font(color=txt_color),
                 number_format=fmt,
                 alignment=_align("right" if c >= 9 else
                                  ("center" if c in (1, 2, 3, 4, 7) else "left")),
                 border=_THIN)
        row += 1

    # Totals
    _total_row(ws, row, 1, 10, label_col=6)
    _set(ws, row, 6, "TOTALS",
         font=_font(bold=True), fill=_fill(_C_GREEN), border=_THIN)
    _set(ws, row, 9, je_summary_data["total_debits"],
         font=_font(bold=True), number_format=_FMT_USD,
         fill=_fill(_C_GREEN), border=_THIN, alignment=_align("right"))
    _set(ws, row, 10, je_summary_data["total_credits"],
         font=_font(bold=True), number_format=_FMT_USD,
         fill=_fill(_C_GREEN), border=_THIN, alignment=_align("right"))
    row += 1

    color = "00B050" if je_summary_data["balanced"] else "FF0000"
    status = "✓  BALANCED" if je_summary_data["balanced"] else "✗  OUT OF BALANCE"
    _set(ws, row, 9, status,
         font=_font(bold=True, color=color), alignment=_align("right"))
    ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=10)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def generate_workpaper(
    period_date: date,
    entity_name: str,
    rates: dict,
    translated_bs: pd.DataFrame,
    translated_is: pd.DataFrame,
    je_lines: list[dict],
    je_summary_data: dict,
    current_col_label: str,
    prior_col_label: str,
    period_label: str,
) -> bytes:
    """
    Build the full workpaper workbook and return it as bytes for download.
    """
    wb = Workbook()
    wb.remove(wb.active)   # Remove the default empty sheet

    _build_summary(wb.create_sheet(),    period_date, entity_name, rates, je_summary_data)
    _build_bs_translation(wb.create_sheet(), translated_bs, current_col_label, prior_col_label)
    _build_is_translation(wb.create_sheet(), translated_is, period_label)
    _build_exch_adj(wb.create_sheet(),       je_lines, rates)
    _build_je_sheet(wb.create_sheet(),       je_lines, je_summary_data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
